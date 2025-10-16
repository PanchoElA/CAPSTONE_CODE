from flask import Flask, request, jsonify, send_file
from flask import render_template
import logging
import sqlite3
import os
import csv
from io import StringIO, BytesIO
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.dimensions import ColumnDimension
    from datetime import datetime
except Exception:
    Workbook = None

app = Flask(__name__)
DB = os.path.join(os.path.dirname(__file__), 'data.db')

def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

# configure logging
logging.basicConfig(level=logging.DEBUG)

def init_db():
    conn = get_db()
    cur = conn.cursor()
    # Tabla original de perfiles/puntos
    cur.execute('''
        CREATE TABLE IF NOT EXISTS profiles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE,
            points INTEGER DEFAULT 0
        )
    ''')
    # Tabla original de scans simples
    cur.execute('''
        CREATE TABLE IF NOT EXISTS scans (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            profile TEXT,
            code TEXT,
            points INTEGER,
            ts INTEGER
        )
    ''')
    # Nueva tabla para QR generados (CODELPA vs TERCERO)
    cur.execute('''
        CREATE TABLE IF NOT EXISTS qr_generated (
            qr_id TEXT PRIMARY KEY,
            marca_envase TEXT,
            origen TEXT,
            peso_envase_kg REAL,
            lote_produccion TEXT,
            fecha_generacion INTEGER,
            tipo_qr TEXT -- 'CODELPA' o 'TERCERO'
        )
    ''')
    # Nueva tabla para retornos REP (datos completos)
    cur.execute('''
        CREATE TABLE IF NOT EXISTS retornos_rep (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            qr_id TEXT,
            marca_envase TEXT,
            origen TEXT,
            fecha_retorno INTEGER,
            estado_retorno TEXT,
            peso_envase_kg REAL,
            destino TEXT, -- 'REUSO_CODELPA' o 'VALORIZACION_INPROPLAS'
            tienda_retorno TEXT,
            evidencia TEXT,
            lote_reporte TEXT,
            cliente_profile TEXT,
            puntos_otorgados INTEGER,
            ts_scan INTEGER,
            FOREIGN KEY(qr_id) REFERENCES qr_generated(qr_id)
        )
    ''')
    conn.commit()
    conn.close()

def startup():
    init_db()

# Ensure DB is initialized when the module is imported/run (safer on minimal Flask installs)
startup()


@app.route('/', methods=['GET'])
def index():
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT name, points FROM profiles ORDER BY name')
    profiles = cur.fetchall()
    cur.execute('SELECT profile, code, points, ts FROM scans ORDER BY ts DESC LIMIT 200')
    scans = cur.fetchall()
    conn.close()
    # convert ts to int (may already be int but sqlite returns it as int)
    scans_list = []
    for s in scans:
        scans_list.append({'profile': s['profile'], 'code': s['code'], 'points': s['points'], 'ts': int(s['ts'] or 0)})
    return render_template('index.html', profiles=profiles, scans=scans_list)


@app.route('/generador', methods=['GET'])
def generador_qr():
    """Página para generar QR únicos"""
    return render_template('generador.html')


@app.route('/reportes', methods=['GET'])
def reportes_rep():
    """Página de reportes REP"""
    import time
    from datetime import datetime
    
    # Últimos 6 meses de lotes
    current_time = datetime.now()
    lotes = []
    for i in range(6):
        if current_time.month - i <= 0:
            year = current_time.year - 1
            month = current_time.month - i + 12
        else:
            year = current_time.year
            month = current_time.month - i
        lotes.append(f"{year}{month:02d}")
    
    return render_template('reportes.html', lotes=lotes)


@app.template_filter('datetimeformat')
def _jinja2_datetimeformat(value):
    try:
        v = int(value)
    except Exception:
        return str(value)
    # handle milliseconds
    if v > 9999999999:
        v = v / 1000.0
    dt = datetime.utcfromtimestamp(v)
    return dt.strftime('%Y-%m-%d %H:%M:%S UTC')

@app.route('/profiles', methods=['GET'])
def list_profiles():
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT name, points FROM profiles')
    rows = cur.fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/points', methods=['POST'])
def add_points():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({'error': 'invalid json'}), 400
    app.logger.debug('POST /points received: %s', data)
    profile = data.get('profile') or 'default'
    points = int(data.get('points', 0))
    code = data.get('code')
    ts = int(data.get('ts', 0))

    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT points FROM profiles WHERE name=?', (profile,))
    row = cur.fetchone()
    if row:
        newp = row['points'] + points
        cur.execute('UPDATE profiles SET points=? WHERE name=?', (newp, profile))
    else:
        cur.execute('INSERT INTO profiles (name, points) VALUES (?, ?)', (profile, points))
    cur.execute('INSERT INTO scans (profile, code, points, ts) VALUES (?, ?, ?, ?)', (profile, code, points, ts))
    conn.commit()
    conn.close()
    # debug: log current profiles table
    conn2 = get_db()
    cur2 = conn2.cursor()
    cur2.execute('SELECT name, points FROM profiles')
    rows2 = cur2.fetchall()
    app.logger.debug('Profiles after update: %s', [dict(r) for r in rows2])
    conn2.close()
    return jsonify({'status':'ok'})


@app.route('/debug/profiles', methods=['GET'])
def debug_profiles():
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM profiles')
    rows = cur.fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/export.csv', methods=['GET'])
def export_csv():
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT name, points FROM profiles')
    rows = cur.fetchall()
    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(['name','points'])
    for r in rows:
        writer.writerow([r['name'], r['points']])
    si.seek(0)
    return send_file(
        BytesIO(si.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name='profiles.csv'
    )


@app.route('/export.xlsx', methods=['GET'])
def export_xlsx():
    if Workbook is None:
        return jsonify({'error': 'openpyxl not installed'}), 500

    conn = get_db()
    cur = conn.cursor()
    # Profiles sheet
    cur.execute('SELECT name, points FROM profiles')
    profiles = cur.fetchall()
    # Scans sheet
    cur.execute('SELECT profile, code, points, ts FROM scans ORDER BY ts')
    scans = cur.fetchall()

    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Profiles'
    headers1 = ['name', 'points']
    ws1.append(headers1)
    # header style
    header_font = Font(bold=True)
    for col_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for r in profiles:
        ws1.append([r['name'], r['points']])

    # adjust column widths for profiles
    for i, col in enumerate(['name', 'points'], 1):
        ws1.column_dimensions[get_column_letter(i)].width = 25 if i == 1 else 12

    ws2 = wb.create_sheet('Scans')
    headers2 = ['profile', 'code', 'points', 'ts']
    ws2.append(headers2)
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for s in scans:
        # convert ts (ms since epoch) to ISO datetime if non-zero
        ts = s['ts']
        if ts:
            try:
                # assume timestamp in milliseconds
                dt = datetime.utcfromtimestamp(int(ts) / 1000.0)
                ts_val = dt.strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                ts_val = str(ts)
        else:
            ts_val = ''
        ws2.append([s['profile'], s['code'], s['points'], ts_val])

    # set column widths and freeze panes
    widths = [18, 40, 10, 20]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='export.xlsx'
    )


# === NUEVOS ENDPOINTS REP ===

@app.route('/generate_qr', methods=['POST'])
def generate_qr():
    """Genera QR único para baldes CODELPA o TERCERO"""
    import uuid
    import time
    
    data = request.get_json(silent=True)
    if not data:
        return jsonify({'error': 'invalid json'}), 400
    
    tipo_qr = data.get('tipo_qr', 'TERCERO')  # 'CODELPA' o 'TERCERO'
    marca_envase = data.get('marca_envase', 'CODELPA' if tipo_qr == 'CODELPA' else 'TERCERO')
    origen = data.get('origen', 'PLANTA_CODELPA' if tipo_qr == 'CODELPA' else 'TIENDA_MP')
    peso_envase_kg = float(data.get('peso_envase_kg', 0.5))
    lote_produccion = data.get('lote_produccion', f'LOTE_{int(time.time())}')
    
    # Generar QR único
    qr_id = f"{tipo_qr}:{uuid.uuid4().hex[:12].upper()}"
    
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute('''
            INSERT INTO qr_generated 
            (qr_id, marca_envase, origen, peso_envase_kg, lote_produccion, fecha_generacion, tipo_qr)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (qr_id, marca_envase, origen, peso_envase_kg, lote_produccion, int(time.time() * 1000), tipo_qr))
        conn.commit()
        
        return jsonify({
            'qr_id': qr_id,
            'marca_envase': marca_envase,
            'origen': origen,
            'peso_envase_kg': peso_envase_kg,
            'lote_produccion': lote_produccion,
            'tipo_qr': tipo_qr
        }), 201  # Código 201 para creación exitosa
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/retorno_rep', methods=['POST'])
def retorno_rep():
    """Registra retorno de balde con datos REP completos - SOLO QRs del sistema desktop"""
    import time
    import re
    
    data = request.get_json(silent=True)
    if not data:
        return jsonify({'error': 'invalid json'}), 400
    
    qr_id = data.get('qr_id')
    if not qr_id:
        return jsonify({'error': 'qr_id requerido'}), 400
    
    # VALIDACIÓN ESTRICTA: Solo QRs del sistema REP desktop
    qr_id = qr_id.strip().upper()
    
    # Validar formato exacto: CODELPA:XXXXXXXXXXXX o TERCERO:XXXXXXXXXXXX
    if not re.match(r'^(CODELPA|TERCERO):[A-F0-9]{12}$', qr_id):
        return jsonify({
            'error': 'QR inválido - Solo se aceptan QRs del sistema REP desktop',
            'formato_requerido': 'CODELPA:XXXXXXXXXXXX o TERCERO:XXXXXXXXXXXX (12 caracteres hexadecimales)',
            'qr_recibido': qr_id
        }), 400
    
    # Buscar QR generado en nuestra base de datos
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM qr_generated WHERE qr_id = ?', (qr_id,))
    qr_info = cur.fetchone()
    
    if not qr_info:
        conn.close()
        return jsonify({
            'error': 'QR no encontrado en el sistema',
            'mensaje': 'Este QR no fue generado por la aplicación desktop del sistema REP',
            'qr_id': qr_id
        }), 404
    
    # Verificar si ya fue procesado
    cur.execute('SELECT retorno_id FROM retornos_rep WHERE qr_id = ?', (qr_id,))
    if cur.fetchone():
        conn.close()
        return jsonify({
            'error': 'QR ya procesado',
            'mensaje': 'Este QR ya fue escaneado y procesado anteriormente',
            'qr_id': qr_id
        }), 409
    
    # Datos del retorno
    estado_retorno = data.get('estado_retorno', 'BUENO')  # BUENO, DAÑADO, REPARABLE
    tienda_retorno = data.get('tienda_retorno', 'MUNDO_PINTURA_CENTRAL')
    cliente_profile = data.get('cliente_profile', 'default')
    evidencia = data.get('evidencia', 'INSPECCION_VISUAL')
    
    # Determinar destino según tipo y estado
    if qr_info['tipo_qr'] == 'CODELPA' and estado_retorno in ['BUENO', 'REPARABLE']:
        destino = 'REUSO_CODELPA'
        puntos_otorgados = 150  # Más puntos por reuso
    else:
        destino = 'VALORIZACION_INPROPLAS'
        puntos_otorgados = 100  # Puntos estándar por valorización
    
    # Generar lote_reporte (YYYYMM)
    from datetime import datetime
    lote_reporte = datetime.now().strftime('%Y%m')
    
    try:
        # Registrar retorno REP
        cur.execute('''
            INSERT INTO retornos_rep 
            (qr_id, marca_envase, origen, fecha_retorno, estado_retorno, peso_envase_kg, 
             destino, tienda_retorno, evidencia, lote_reporte, cliente_profile, puntos_otorgados, ts_scan)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (qr_id, qr_info['marca_envase'], qr_info['origen'], int(time.time() * 1000), 
              estado_retorno, qr_info['peso_envase_kg'], destino, tienda_retorno, 
              evidencia, lote_reporte, cliente_profile, puntos_otorgados, int(time.time() * 1000)))
        
        # Actualizar puntos del cliente
        cur.execute('SELECT points FROM profiles WHERE name=?', (cliente_profile,))
        row = cur.fetchone()
        if row:
            newp = row['points'] + puntos_otorgados
            cur.execute('UPDATE profiles SET points=? WHERE name=?', (newp, cliente_profile))
        else:
            cur.execute('INSERT INTO profiles (name, points) VALUES (?, ?)', (cliente_profile, puntos_otorgados))
        
        conn.commit()
        
        app.logger.debug(f'Retorno REP: {qr_id} -> {destino}, puntos: {puntos_otorgados}')
        
        return jsonify({
            'status': 'ok',
            'qr_id': qr_id,
            'destino': destino,
            'puntos_otorgados': puntos_otorgados,
            'lote_reporte': lote_reporte
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/reporte_rep/<lote>', methods=['GET'])
def reporte_rep(lote):
    """Genera reporte REP por lote (YYYYMM)"""
    conn = get_db()
    cur = conn.cursor()
    
    # Retornos del lote
    cur.execute('''
        SELECT qr_id, marca_envase, origen, fecha_retorno, estado_retorno, 
               peso_envase_kg, destino, tienda_retorno, evidencia, lote_reporte,
               cliente_profile, puntos_otorgados
        FROM retornos_rep 
        WHERE lote_reporte = ?
        ORDER BY fecha_retorno
    ''', (lote,))
    
    retornos = [dict(r) for r in cur.fetchall()]
    
    # Estadísticas
    reuso_count = sum(1 for r in retornos if r['destino'] == 'REUSO_CODELPA')
    valorizacion_count = sum(1 for r in retornos if r['destino'] == 'VALORIZACION_INPROPLAS')
    peso_total_reuso = sum(r['peso_envase_kg'] for r in retornos if r['destino'] == 'REUSO_CODELPA')
    peso_total_valorizacion = sum(r['peso_envase_kg'] for r in retornos if r['destino'] == 'VALORIZACION_INPROPLAS')
    
    conn.close()
    
    return jsonify({
        'lote_reporte': lote,
        'resumen': {
            'total_retornos': len(retornos),
            'reuso_codelpa': reuso_count,
            'valorizacion_inproplas': valorizacion_count,
            'peso_total_reuso_kg': peso_total_reuso,
            'peso_total_valorizacion_kg': peso_total_valorizacion
        },
        'retornos': retornos
    })


@app.route('/cliente_retornos/<cliente>', methods=['GET'])
def cliente_retornos(cliente):
    """Obtiene resumen completo de un cliente específico"""
    conn = get_db()
    cur = conn.cursor()
    
    # Obtener perfil del cliente
    cur.execute('SELECT name, points FROM profiles WHERE name = ? COLLATE NOCASE', (cliente,))
    profile_row = cur.fetchone()
    
    if not profile_row:
        return jsonify({"error": "Cliente no encontrado"}), 404
    
    profile = {"name": profile_row[0], "points": profile_row[1]}
    
    # Obtener retornos recientes (últimos 10)
    cur.execute('''
        SELECT qr_id, marca_envase, fecha_retorno, estado_retorno, 
               peso_envase_kg, destino, puntos_otorgados, tienda_retorno
        FROM retornos_rep 
        WHERE cliente_profile = ? COLLATE NOCASE
        ORDER BY fecha_retorno DESC
        LIMIT 10
    ''', (cliente,))
    
    retornos_recientes = [dict(r) for r in cur.fetchall()]
    
    # Obtener totales
    cur.execute('''
        SELECT COUNT(*), COALESCE(SUM(peso_envase_kg), 0)
        FROM retornos_rep 
        WHERE cliente_profile = ? COLLATE NOCASE
    ''', (cliente,))
    
    totales = cur.fetchone()
    total_retornos = totales[0] if totales else 0
    peso_total_kg = totales[1] if totales else 0.0
    
    conn.close()
    
    resumen = {
        "profile": profile,
        "retornos_recientes": retornos_recientes,
        "total_retornos": total_retornos,
        "peso_total_kg": peso_total_kg
    }
    
    return jsonify(resumen)

@app.route('/retornos_completos/<cliente>', methods=['GET'])
def retornos_completos(cliente):
    """Obtiene historial completo de retornos de un cliente"""
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute('''
        SELECT qr_id, marca_envase, fecha_retorno, estado_retorno, 
               peso_envase_kg, destino, puntos_otorgados, tienda_retorno
        FROM retornos_rep 
        WHERE cliente_profile = ? COLLATE NOCASE
        ORDER BY fecha_retorno DESC
    ''', (cliente,))
    
    retornos = [dict(r) for r in cur.fetchall()]
    conn.close()
    
    return jsonify(retornos)

@app.route('/stats', methods=['GET'])
def get_stats():
    """Obtiene estadísticas generales del sistema REP"""
    conn = get_db()
    cur = conn.cursor()
    
    # QRs generados
    cur.execute('SELECT COUNT(*) FROM qr_generated')
    qrs_generados = cur.fetchone()[0]
    
    # Total retornos
    cur.execute('SELECT COUNT(*) FROM retornos_rep')
    retornos_total = cur.fetchone()[0]
    
    # Peso total
    cur.execute('SELECT COALESCE(SUM(peso_envase_kg), 0) FROM retornos_rep')
    peso_total_kg = cur.fetchone()[0]
    
    # Puntos totales otorgados
    cur.execute('SELECT COALESCE(SUM(puntos_otorgados), 0) FROM retornos_rep')
    puntos_total = cur.fetchone()[0]
    
    # Estadísticas por destino
    cur.execute('''
        SELECT destino, COUNT(*) 
        FROM retornos_rep 
        GROUP BY destino
    ''')
    destinos = dict(cur.fetchall())
    
    # Estadísticas por estado
    cur.execute('''
        SELECT estado_retorno, COUNT(*) 
        FROM retornos_rep 
        GROUP BY estado_retorno
    ''')
    estados = dict(cur.fetchall())
    
    conn.close()
    
    stats = {
        'qrs_generados': qrs_generados,
        'retornos_total': retornos_total,
        'peso_total_kg': float(peso_total_kg),
        'puntos_total': puntos_total,
        'reuso_codelpa': destinos.get('REUSO_CODELPA', 0),
        'valorizacion_inproplas': destinos.get('VALORIZACION_INPROPLAS', 0),
        'estado_bueno': estados.get('BUENO', 0),
        'estado_reparable': estados.get('REPARABLE', 0),
        'estado_danado': estados.get('DAÑADO', 0)
    }
    
    return jsonify(stats)


@app.route('/export_rep_csv/<lote>', methods=['GET'])
def export_rep_csv(lote):
    """Exporta CSV de retornos REP por lote"""
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute('''
        SELECT qr_id, marca_envase, origen, fecha_retorno, estado_retorno, 
               peso_envase_kg, destino, tienda_retorno, evidencia, lote_reporte
        FROM retornos_rep 
        WHERE lote_reporte = ?
        ORDER BY fecha_retorno
    ''', (lote,))
    
    retornos = cur.fetchall()
    conn.close()
    
    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(['qr_id', 'marca_envase', 'origen', 'fecha_retorno', 'estado_retorno', 
                     'peso_envase_kg', 'destino', 'tienda_retorno', 'evidencia', 'lote_reporte'])
    
    for r in retornos:
        writer.writerow([r['qr_id'], r['marca_envase'], r['origen'], r['fecha_retorno'], 
                        r['estado_retorno'], r['peso_envase_kg'], r['destino'], 
                        r['tienda_retorno'], r['evidencia'], r['lote_reporte']])
    
    si.seek(0)
    return send_file(
        BytesIO(si.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'reporte_rep_{lote}.csv'
    )


@app.route('/demo/generar_datos', methods=['POST'])
def generar_datos_demo():
    """Genera datos de ejemplo para demostrar el sistema REP"""
    import uuid
    import time
    import random
    from datetime import datetime, timedelta
    
    conn = get_db()
    cur = conn.cursor()
    
    # Generar QRs de ejemplo
    qrs_generados = []
    marcas_codelpa = ["CODELPA", "SHERWIN_WILLIAMS"]
    marcas_tercero = ["ANYPSA", "CHILCORROFIN", "COMEX"]
    
    # 20 QRs CODELPA
    for i in range(20):
        qr_id = f"CODELPA:{uuid.uuid4().hex[:12].upper()}"
        marca = random.choice(marcas_codelpa)
        peso = round(random.uniform(0.4, 0.8), 2)
        
        cur.execute('''
            INSERT OR IGNORE INTO qr_generated 
            (qr_id, marca_envase, origen, peso_envase_kg, lote_produccion, fecha_generacion, tipo_qr)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (qr_id, marca, "PLANTA_CODELPA", peso, f"LOTE_2024_10_{i:03d}", int(time.time() * 1000), "CODELPA"))
        qrs_generados.append(qr_id)
    
    # 15 QRs TERCERO
    for i in range(15):
        qr_id = f"TERCERO:{uuid.uuid4().hex[:12].upper()}"
        marca = random.choice(marcas_tercero)
        peso = round(random.uniform(0.3, 0.6), 2)
        
        cur.execute('''
            INSERT OR IGNORE INTO qr_generated 
            (qr_id, marca_envase, origen, peso_envase_kg, lote_produccion, fecha_generacion, tipo_qr)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (qr_id, marca, "TIENDA_MP", peso, f"LOTE_EXT_{i:03d}", int(time.time() * 1000), "TERCERO"))
        qrs_generados.append(qr_id)
    
    # Simular retornos durante los últimos 3 meses
    clientes = ["Juan Pérez", "María García", "Carlos López", "Ana Silva", "Pedro Morales"]
    estados = ["BUENO", "DAÑADO", "REPARABLE"]
    tiendas = ["MUNDO_PINTURA_CENTRAL", "MUNDO_PINTURA_NORTE", "MUNDO_PINTURA_SUR"]
    
    retornos_creados = 0
    for qr_id in qrs_generados[:25]:  # Solo 25 de los 35 QRs han sido devueltos
        # Fecha aleatoria en últimos 90 días
        dias_atras = random.randint(1, 90)
        fecha_retorno = datetime.now() - timedelta(days=dias_atras)
        lote_reporte = fecha_retorno.strftime('%Y%m')
        
        # Buscar info del QR
        cur.execute('SELECT * FROM qr_generated WHERE qr_id = ?', (qr_id,))
        qr_info = cur.fetchone()
        
        estado = random.choice(estados)
        cliente = random.choice(clientes)
        tienda = random.choice(tiendas)
        
        # Determinar destino y puntos
        if qr_info['tipo_qr'] == 'CODELPA' and estado in ['BUENO', 'REPARABLE']:
            destino = 'REUSO_CODELPA'
            puntos = 150
        else:
            destino = 'VALORIZACION_INPROPLAS'
            puntos = 100
        
        cur.execute('''
            INSERT OR IGNORE INTO retornos_rep 
            (qr_id, marca_envase, origen, fecha_retorno, estado_retorno, peso_envase_kg, 
             destino, tienda_retorno, evidencia, lote_reporte, cliente_profile, puntos_otorgados, ts_scan)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (qr_id, qr_info['marca_envase'], qr_info['origen'], int(fecha_retorno.timestamp() * 1000),
              estado, qr_info['peso_envase_kg'], destino, tienda, "INSPECCION_VISUAL", 
              lote_reporte, cliente, puntos, int(fecha_retorno.timestamp() * 1000)))
        
        # Actualizar puntos del cliente
        cur.execute('SELECT points FROM profiles WHERE name=?', (cliente,))
        row = cur.fetchone()
        if row:
            nuevos_puntos = row['points'] + puntos
            cur.execute('UPDATE profiles SET points=? WHERE name=?', (nuevos_puntos, cliente))
        else:
            cur.execute('INSERT INTO profiles (name, points) VALUES (?, ?)', (cliente, puntos))
        
        retornos_creados += 1
    
    conn.commit()
    conn.close()
    
    return jsonify({
        'status': 'ok',
        'qrs_generados': len(qrs_generados),
        'retornos_creados': retornos_creados,
        'message': 'Datos de demo generados exitosamente'
    })


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
